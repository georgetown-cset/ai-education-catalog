"""
Reformat xlsx export of AI Education Catalog google sheet into a CSV
"""
import argparse
import csv
import os

from openpyxl import load_workbook


def get_rows(filename: str) -> iter:
    """
    Reads xlsx file and returns generator of cleaned rows
    :param filename: xlsx filename
    :return: Generator of cleaned rows
    """
    name_to_comment_rows = {
        "After School Programs": 4,
        "Conferences Challenges": 2,
        "Curriculum": 3,
        "Federal Initiatives": 2,
        "Scholarships": 2,
        "Summer Camps": 3
    }
    wb = load_workbook(filename)
    for sheet in wb:
        if sheet.title not in name_to_comment_rows:
            print("Skipping "+sheet.title)
            continue
        num_commment_rows = name_to_comment_rows[sheet.title]
        col_names = None
        for row in sheet.iter_rows(min_row=num_commment_rows+1):
            if col_names is None:
                col_names = [c.value.strip() for c in row if c.value is not None]
                assert col_names[0] == "Program", f"Expected first column to be 'Program' in {sheet.title}"
            else:
                stripped_row = {key: "" if row[idx].value is None else str(row[idx].value).strip().strip(",").strip()
                             for idx, key in enumerate(col_names)}
                if not stripped_row["Program"]:
                    continue
                stripped_row["URL"] = row[0].hyperlink.target
                yield stripped_row


def reformat_sheet(input_sheet: str, output_csv: str):
    """
    Consolidates xlsx export of CSET google sheet
    (https://docs.google.com/spreadsheets/d/1y-Ez9NY1nhSyewMOqbosGueSqieiFinj/edit#gid=2077629231)
    into a single csv with only columns that are relevant to the UI
    :param input_sheet: xlsx download of google sheet
    :param output_csv: path to csv where output should be written
    :return: None
    """
    rows = get_rows(input_sheet)
    output_cols = ["Community", "Cost", "Detailed Location", "Duration", "Gender", "Level", "Location",
                   "Objective", "Organization Type", "Pre-recs", "Program", "Race/Ethnicity", "Target",
                   "Type", "URL", "Underrepresented"]
    with open(output_csv, mode="w") as f:
        writer = csv.DictWriter(f, fieldnames=output_cols)
        writer.writeheader()
        for row in rows:
            relevant_row = {k: v for k, v in row.items() if k in output_cols}
            writer.writerow(relevant_row)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input_sheet")
    parser.add_argument("--output_csv", default=os.path.join("raw_data", "ai_education_catalog.csv"))
    args = parser.parse_args()

    reformat_sheet(args.input_sheet, args.output_csv)