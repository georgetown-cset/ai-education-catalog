"""
Reformat AI Education Catalog spreadsheet into javascript object.
"""
import argparse
import json
import os
import re
import us

from openpyxl import load_workbook


EXPECTED_TYPES = {"After-School Program", "Apprenticeship", "Challenge", "Conference", "Curriculum", "Fellowship",
                      "Hackathon", "Internship", "Robotics", "Scholarship", "Summer Camp"}
EXPECTED_ORGANIZATIONS = {"Government", "Non-Profit", "Private", "Public", "University"}
EXPECTED_TARGET_AUDIENCE = {"Anyone", "Educators", "Elementary school students", "Middle school students",
                            "High school students", "Postsecondary students", "Professionals"}
REQUIRED_KEYS = ["name", "url", "type", "organization", "location", "type", "target"]


def reformat_data(input_fi: str) -> list:
    """
    Reformat xlsx catalog download from
    https://docs.google.com/spreadsheets/d/1y-Ez9NY1nhSyewMOqbosGueSqieiFinj/edit#gid=2077629231
    into list of cleaned rows
    :param input_fi: path to the xlsx download
    :return: list of cleaned rows
    """
    cleaned_data = []
    missing_location = []
    # get_rows returns an iterator of dicts that have had their keys and values stripped
    for counter, line in enumerate(get_rows(input_fi)):
        orig_locations = line.get("Location")
        if orig_locations:
            orig_locations = orig_locations.replace("USA", "National")
        locations = clean_locations(orig_locations, line["Program"])
        if not locations:
            missing_location.append(line["Program"]+"-"+line["Type"])
            locations = ["National"]
        targets = get_targets(line.get("Target"))
        pre_reqs = [pr.strip() for pr in line.get("Pre-recs", "").split(",") if len(pr.strip()) > 0]
        short_obj = get_short_objective(line.get("Objective"))
        pre_check_row(line)
        row = {
            "id": counter,
            "name": line["Program"],
            "url": line["URL"],
            "type": line["Type"],
            "organization": line.get("Organization Type"),
            "target": targets,
            "is_free": line.get("Cost", "").strip().lower() == "free",
            "location": locations,
            "is_not_virtual": "Virtual" not in locations,
            "is_not_national": ("National" not in locations) and (len(locations) > 0),
            "location_details": get_detailed_location(orig_locations, line.get("Detailed Location")),
            "is_underrep": bool(line.get("Underrepresented")),
            "gender": [g.strip().title() for g in line.get("Gender", "").split(",")],
            "race_ethnicity": [g.strip().title() for g in line.get("Race/Ethnicity", "").split(",")],
            "is_community_program": bool(line.get("Community")),
            "objective": line.get("Objective"),
            "short_objective": short_obj,
            "level": line.get("Level"),
            "cost": clean_cost(line.get("Cost")),
            "pre_reqs": pre_reqs,
            "duration": line.get("Duration")
        }
        clean_row = postprocess_row(row)
        cleaned_data.append(clean_row)
    cleaned_data.sort(key=lambda r: r["name"])
    print(f"Missing location for {missing_location}, replaced with 'National'")
    return cleaned_data


def pre_check_row(line: dict) -> None:
    """
    Runs data checks and prints or raises error depending on severity
    :param line: Line of data
    :return: None
    """
    row_id = f"'{line['Program']}' in {line['Type']}"
    if line["Type"].title() not in EXPECTED_TYPES:
        print(f"Unexpected program type for {row_id}")
    if line.get("Level") and "00:00" in line.get("Level"):
        raise ValueError(f"Level has timestamp type in {row_id}")
    if line.get("Organization Type") not in EXPECTED_ORGANIZATIONS:
        print(f"Unexpected organization type for {row_id}.")


def postprocess_row(row: dict) -> dict:
    """
    Does shared post-processing of data values and final checks
    :param row: row of data
    :return: cleaned row
    """
    clean_row = {}
    for k, v in row.items():
        if type(v) == str:
            v = v.strip()
            if k not in {"name", "objective", "short_objective", "level", "url", "location_details"}:
                v = v.title()
            if not v:
                v = None
        elif type(v) == list:
            v = [elt.strip() for elt in v if len(elt.strip()) > 0]
        clean_row[k] = v
    for key in REQUIRED_KEYS:
        if not clean_row[key]:
            print(f"Empty value for {key} in '{row['name']}' in {row['type']}")
    return clean_row


def get_detailed_location(general_locations: list, detailed_location: str) -> str:
    """
    Takes general locations and detailed location and returns a location string that is as detailed
    as possible but does not contain redundant information
    :param general_locations: List of high-level locations (like Oklahoma, Texas, Virtual)
    :param detailed_location: A detailed location (like "Oklahoma City, Oklahoma") or None if the
        detailed location is completely redundant
    :return: normalized detailed location
    """
    if not detailed_location:
        return None
    clean_detailed_location = detailed_location.strip().strip("*").strip()
    if re.search(r"^[a-z]", detailed_location):
        # only titlecase if we think it's lowercase because some detailed locations start
        # with acronyms (MIT, NYU...)
        clean_detailed_location = clean_detailed_location.title()
    if detailed_location in general_locations:
        return None
    return clean_detailed_location


def clean_cost(cost: str) -> str:
    """
    Cleans program cost, reformatting to a dollar sign followed by numbers if possible,
    and normalizing various unspecified variants to "Cost Not Specified"
    :param cost: string representing cost of the program
    :return: normalized cost string
    """
    if not cost:
        return "Cost Not Specified"
    cost = cost.strip()
    if re.search(r"\d", cost):
        try:
            cost = str(int(float(cost.strip("$").strip())))
        except:
            # not an issue if the above works out, it's just an attempt to normalize strings like
            # $100.00 into $100
            pass
        if re.search(r"^\d", cost):
            return f"${cost}"
    elif cost.lower() in {"request a quote", "not specified"}:
        return "Cost Not Specified"
    return cost


def get_short_objective(objective: str) -> str:
    """
    Truncates objective to approximately 470 characters
    :param objective: Program objective description
    :return: Shortened program objective
    """
    soft_char_limit = 440
    if objective is None or len(objective) <= soft_char_limit:
        return objective
    words = objective.split()
    short_objective = ""
    for idx, word in enumerate(words):
        if len(short_objective) >= soft_char_limit:
            break
        short_objective += word + " "
    return short_objective.strip()+"..."


def get_targets(raw_targets: str) -> list:
    """
    Cleans list of program target participants
    :param raw_targets: list of program target participants
    :return: list of cleaned target participants
    """
    targets = [t.strip().title() for t in raw_targets.replace(".", ",").split(",")]
    clean_targets = []
    for target in targets:
        if not target:
            continue
        clean_target = target
        if target in {"Elementary", "Middle", "High"}:
            clean_target = target+" school students"
        elif target == "Postsecondary":
            clean_target = target+" students"
        assert clean_target in EXPECTED_TARGET_AUDIENCE, clean_target
        clean_targets.append(clean_target)
    return clean_targets


def clean_locations(location: str, program_name: str) -> list:
    """
    Cleans locations, including turning two-character state names into full names.
    :param location: string representing location
    :param program_name: name of the program (used in warning messages)
    :return: cleaned/expanded location
    """
    expected_locations = ["Global", "National", "Virtual", "District of Columbia"] + \
                         [s.name for s in us.states.STATES_AND_TERRITORIES]
    if not location:
        return []
    clean_locations = set()
    for loc in location.strip().split(","):
        loc = loc.strip()
        if len(loc) == 2:
            loc_obj = us.states.lookup(loc)
            if loc_obj is None:
                print(f"Could not convert location for '{program_name}': {loc}")
            else:
                loc = loc_obj.name.title().replace(" Of ", " of ")
        clean_locations.add(loc)
        if loc not in expected_locations:
            print(f"Unexpected location for '{program_name}': {loc}")
    return list(clean_locations)


def get_rows(filename: str) -> iter:
    """
    Reads xlsx file and returns generator of cleaned rows
    :param filename: xlsx filename
    :return: Generator of cleaned rows
    """
    name_to_comment_rows = {
        "After School Programs": 4,
        "Conferences Challenges": 2,
        "Curriculum": 3,
        "Federal Initiatives": 2,
        "Scholarships": 2,
        "Summer Camps": 3
    }
    wb = load_workbook(filename)
    for sheet in wb:
        if sheet.title not in name_to_comment_rows:
            print("Skipping "+sheet.title)
            continue
        num_commment_rows = name_to_comment_rows[sheet.title]
        col_names = None
        for row in sheet.iter_rows(min_row=num_commment_rows+1):
            if col_names is None:
                col_names = [c.value.strip() for c in row if c.value is not None]
                assert col_names[0] == "Program", f"Expected first column to be 'Program' in {sheet.title}"
            else:
                stripped_row = {key: "" if row[idx].value is None else str(row[idx].value).strip().strip(",").strip()
                             for idx, key in enumerate(col_names)}
                if not stripped_row["Program"]:
                    continue
                stripped_row["URL"] = row[0].hyperlink.target
                yield stripped_row


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw_data", default="raw_data/AI Education Catalog.xlsx")
    parser.add_argument("--output_dir", default="../ai-education-catalog/src/data")
    args = parser.parse_args()

    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)

    cleaned_data = reformat_data(args.raw_data)
    with open(os.path.join(args.output_dir, "data.js"), mode="w") as f:
        f.write("const data = "+json.dumps(cleaned_data)+"\n\n\nexport {data};")
